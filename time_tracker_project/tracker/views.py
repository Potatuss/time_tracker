from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q, Sum
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect, render
from datetime import datetime, date, timedelta
import json
import calendar
from .models import EntradaTiempo, Cliente, Proyecto, Etiqueta
from .forms import EntradaTiempoForm, ClienteForm, ProyectoForm, ReporteForm
from django.views.decorators.csrf import csrf_exempt

def dashboard(request):
    """Vista principal del dashboard"""
    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_mes = hoy.replace(day=1)
    
    # Estadísticas rápidas
    stats = {
        'horas_hoy': EntradaTiempo.objects.filter(fecha=hoy).aggregate(
            total=Sum('horas'))['total'] or 0,
        'horas_semana': EntradaTiempo.objects.filter(fecha__gte=inicio_semana).aggregate(
            total=Sum('horas'))['total'] or 0,
        'horas_mes': EntradaTiempo.objects.filter(fecha__gte=inicio_mes).aggregate(
            total=Sum('horas'))['total'] or 0,
        'proyectos_activos': Proyecto.objects.filter(estado='activo').count(),
    }
    
    # Entradas recientes
    entradas_recientes = EntradaTiempo.objects.select_related('proyecto', 'proyecto__cliente').order_by('-fecha', '-hora_inicio')[:5]
    
    # Datos para el modal
    proyectos = Proyecto.objects.filter(estado='activo').select_related('cliente')
    etiquetas = Etiqueta.objects.all()
    clientes = Cliente.objects.filter(activo=True)
    
    context = {
        'stats': stats,
        'entradas_recientes': entradas_recientes,
        'proyectos': proyectos,
        'etiquetas': etiquetas,
        'clientes': clientes,
    }
    
    return render(request, 'tracker/dashboard.html', context)

def calendario_mensual(request):
    """Vista del calendario mensual"""
    year = int(request.GET.get('year', date.today().year))
    month = int(request.GET.get('month', date.today().month))
    
    # Navegar entre meses
    fecha_actual = date(year, month, 1)
    if month == 1:
        mes_anterior = fecha_actual.replace(year=year-1, month=12)
    else:
        mes_anterior = fecha_actual.replace(month=month-1)
    
    if month == 12:
        mes_siguiente = fecha_actual.replace(year=year+1, month=1)
    else:
        mes_siguiente = fecha_actual.replace(month=month+1)
    
    # Obtener entradas del mes
    primer_dia = date(year, month, 1)
    ultimo_dia = date(year, month, calendar.monthrange(year, month)[1])
    
    entradas = EntradaTiempo.objects.filter(
        fecha__gte=primer_dia,
        fecha__lte=ultimo_dia
    ).select_related('proyecto', 'proyecto__cliente').prefetch_related('etiquetas').order_by('-fecha', '-hora_inicio')
    
    # Calcular estadísticas del mes
    total_horas_mes = sum(entrada.horas for entrada in entradas)
    dias_trabajados = len(set(entrada.fecha for entrada in entradas))
    proyectos_unicos = len(set(entrada.proyecto for entrada in entradas))
    promedio_diario = total_horas_mes / dias_trabajados if dias_trabajados > 0 else 0
    total_facturado = sum(
        entrada.horas * (entrada.proyecto.tarifa_por_hora or 0) 
        for entrada in entradas
    )
    
    context = {
        'year': year,
        'month': month,
        'fecha_actual': fecha_actual,
        'mes_anterior': mes_anterior,
        'mes_siguiente': mes_siguiente,
        'entradas': entradas,
        'nombre_mes': calendar.month_name[month],
        'total_horas_mes': total_horas_mes,
        'dias_trabajados': dias_trabajados,
        'proyectos_unicos': proyectos_unicos,
        'promedio_diario': promedio_diario,
        'total_facturado': total_facturado,
    }
    
    return render(request, 'tracker/calendario_mensual.html', context)

def calendario_semanal(request):
    """Vista del calendario semanal"""
    # Obtener fecha de inicio de semana
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        fecha_base = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha_base = date.today()
    
    # Calcular inicio de semana (lunes)
    inicio_semana = fecha_base - timedelta(days=fecha_base.weekday())
    fin_semana = inicio_semana + timedelta(days=6)
    
    # Semanas anterior y siguiente
    semana_anterior = inicio_semana - timedelta(days=7)
    semana_siguiente = inicio_semana + timedelta(days=7)
    
    # Crear lista de días de la semana
    dias_semana = []
    total_horas_semana = 0
    
    for i in range(7):
        dia = inicio_semana + timedelta(days=i)
        entradas_dia = EntradaTiempo.objects.filter(fecha=dia).select_related('proyecto', 'proyecto__cliente').prefetch_related('etiquetas').order_by('hora_inicio')
        total_horas_dia = sum(entrada.horas for entrada in entradas_dia)
        total_horas_semana += total_horas_dia
        
        dias_semana.append({
            'fecha': dia,
            'entradas': entradas_dia,
            'total_horas': total_horas_dia,
        })
    
    context = {
        'inicio_semana': inicio_semana,
        'fin_semana': fin_semana,
        'semana_anterior': semana_anterior,
        'semana_siguiente': semana_siguiente,
        'dias_semana': dias_semana,
        'total_horas': total_horas_semana,
    }
    
    return render(request, 'tracker/calendario_semanal.html', context)

def calendario_diario(request):
    """Vista del calendario diario"""
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha = date.today()
    
    dia_anterior = fecha - timedelta(days=1)
    dia_siguiente = fecha + timedelta(days=1)
    
    # Entradas del día
    entradas = EntradaTiempo.objects.filter(fecha=fecha).select_related('proyecto', 'proyecto__cliente').prefetch_related('etiquetas').order_by('hora_inicio')
    
    total_horas = sum(entrada.horas for entrada in entradas)
    
    context = {
        'fecha': fecha,
        'dia_anterior': dia_anterior,
        'dia_siguiente': dia_siguiente,
        'entradas': entradas,
        'total_horas': total_horas,
    }
    
    return render(request, 'tracker/calendario_diario.html', context)

def nueva_entrada(request):
    """Crear nueva entrada de tiempo"""
    if request.method == 'POST':
        form = EntradaTiempoForm(request.POST, user=request.user)
        if form.is_valid():
            entrada = form.save(commit=False)
            entrada.usuario = request.user
            entrada.save()
            form.save_m2m()  # Guardar relaciones many-to-many
            messages.success(request, 'Entrada de tiempo agregada exitosamente.')
            return redirect('tracker:dashboard')
    else:
        form = EntradaTiempoForm(user=request.user)
        # Pre-llenar fecha si viene como parámetro
        fecha_param = request.GET.get('fecha')
        if fecha_param:
            try:
                fecha = datetime.strptime(fecha_param, '%Y-%m-%d').date()
                form.fields['fecha'].initial = fecha
            except ValueError:
                pass
    
    context = {
        'form': form,
        'titulo': 'Nueva Entrada de Tiempo'
    }
    
    return render(request, 'tracker/entrada_form.html', context)

def editar_entrada(request, pk):
    """Editar entrada de tiempo existente"""
    entrada = get_object_or_404(EntradaTiempo, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = EntradaTiempoForm(request.POST, instance=entrada, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Entrada de tiempo actualizada exitosamente.')
            return redirect('tracker:dashboard')
    else:
        form = EntradaTiempoForm(instance=entrada, user=request.user)
    
    context = {
        'form': form,
        'entrada': entrada,
        'titulo': 'Editar Entrada de Tiempo'
    }
    
    return render(request, 'tracker/entrada_form.html', context)

def eliminar_entrada(request, pk):
    """Eliminar entrada de tiempo"""
    entrada = get_object_or_404(EntradaTiempo, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        entrada.delete()
        messages.success(request, 'Entrada de tiempo eliminada exitosamente.')
        return redirect('tracker:dashboard')
    
    return render(request, 'tracker/confirmar_eliminacion.html', {'entrada': entrada})

def entradas_json(request):
    """API JSON para entradas de tiempo (para calendario)"""
    try:
        fecha_inicio = request.GET.get('start')
        fecha_fin = request.GET.get('end')
        
        entradas = EntradaTiempo.objects.all()
        
        # Parsear fechas si están presentes
        if fecha_inicio:
            # Convertir de ISO string a fecha
            fecha_inicio_parsed = datetime.fromisoformat(fecha_inicio.replace('Z', '+00:00')).date()
            entradas = entradas.filter(fecha__gte=fecha_inicio_parsed)
        if fecha_fin:
            fecha_fin_parsed = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00')).date()
            entradas = entradas.filter(fecha__lte=fecha_fin_parsed)
        
        eventos = []
        for entrada in entradas.select_related('proyecto', 'proyecto__cliente'):
            # Crear eventos para FullCalendar
            evento = {
                'id': entrada.id,
                'title': f"{entrada.proyecto.nombre} ({entrada.horas}h)",
                'start': f"{entrada.fecha}T{entrada.hora_inicio}",
                'description': entrada.descripcion,
                'backgroundColor': '#2563eb',
                'borderColor': '#2563eb',
                'extendedProps': {
                    'proyecto': entrada.proyecto.nombre,
                    'proyecto_id': entrada.proyecto.id,
                    'cliente': entrada.proyecto.cliente.nombre,
                    'horas': str(entrada.horas),
                    'descripcion': entrada.descripcion,
                    'etiquetas': [etiqueta.nombre for etiqueta in entrada.etiquetas.all()],
                }
            }
            
            # Agregar hora de fin si existe
            if entrada.hora_fin:
                evento['end'] = f"{entrada.fecha}T{entrada.hora_fin}"
            
            eventos.append(evento)
        
        return JsonResponse(eventos, safe=False)
        
    except Exception as e:
        # En caso de error, devolver lista vacía
        return JsonResponse([], safe=False)

def entrada_detalle_json(request, pk):
    """Detalle de entrada en JSON"""
    entrada = get_object_or_404(EntradaTiempo, pk=pk)
    
    data = {
        'id': entrada.id,
        'proyecto': entrada.proyecto.nombre,
        'proyecto_id': entrada.proyecto.id,
        'cliente': entrada.proyecto.cliente.nombre,
        'fecha': entrada.fecha.strftime('%Y-%m-%d'),
        'hora_inicio': entrada.hora_inicio.strftime('%H:%M'),
        'hora_fin': entrada.hora_fin.strftime('%H:%M') if entrada.hora_fin else '',
        'horas': str(entrada.horas),
        'descripcion': entrada.descripcion,
        'etiquetas': [etiqueta.nombre for etiqueta in entrada.etiquetas.all()],
        'etiquetas_ids': [etiqueta.id for etiqueta in entrada.etiquetas.all()],
    }
    
    return JsonResponse(data)

def nuevo_cliente(request):
    """Crear nuevo cliente o listar clientes existentes"""
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente creado exitosamente.')
            return redirect('tracker:nuevo_cliente')
    else:
        form = ClienteForm()
    
    context = {
        'form': form,
        'clientes': clientes
    }
    
    return render(request, 'tracker/clientes.html', context)

def nuevo_proyecto(request):
    """Crear nuevo proyecto o listar proyectos existentes"""
    proyectos = Proyecto.objects.filter(estado='activo').select_related('cliente').order_by('-fecha_creacion')
    
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proyecto creado exitosamente.')
            return redirect('tracker:nuevo_proyecto')
    else:
        form = ProyectoForm()
    
    context = {
        'form': form,
        'proyectos': proyectos
    }
    
    return render(request, 'tracker/proyectos.html', context)

def reportes(request):
    """Vista de reportes"""
    form = ReporteForm()
    entradas = []
    total_horas = 0
    total_facturado = 0
    estadisticas = {}
    
    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            # Obtener todas las entradas
            entradas = EntradaTiempo.objects.all().select_related('proyecto', 'proyecto__cliente').prefetch_related('etiquetas')
            
            # Aplicar filtros de fecha según el período
            periodo = form.cleaned_data['periodo']
            hoy = date.today()
            
            if periodo == 'hoy':
                entradas = entradas.filter(fecha=hoy)
            elif periodo == 'ayer':
                ayer = hoy - timedelta(days=1)
                entradas = entradas.filter(fecha=ayer)
            elif periodo == 'esta_semana':
                inicio_semana = hoy - timedelta(days=hoy.weekday())
                entradas = entradas.filter(fecha__gte=inicio_semana)
            elif periodo == 'semana_pasada':
                inicio_semana_pasada = hoy - timedelta(days=hoy.weekday() + 7)
                fin_semana_pasada = inicio_semana_pasada + timedelta(days=6)
                entradas = entradas.filter(fecha__gte=inicio_semana_pasada, fecha__lte=fin_semana_pasada)
            elif periodo == 'este_mes':
                inicio_mes = hoy.replace(day=1)
                entradas = entradas.filter(fecha__gte=inicio_mes)
            elif periodo == 'mes_pasado':
                if hoy.month == 1:
                    mes_pasado = hoy.replace(year=hoy.year-1, month=12, day=1)
                else:
                    mes_pasado = hoy.replace(month=hoy.month-1, day=1)
                
                # Último día del mes pasado
                if mes_pasado.month == 12:
                    fin_mes_pasado = mes_pasado.replace(year=mes_pasado.year+1, month=1, day=1) - timedelta(days=1)
                else:
                    fin_mes_pasado = mes_pasado.replace(month=mes_pasado.month+1, day=1) - timedelta(days=1)
                
                entradas = entradas.filter(fecha__gte=mes_pasado, fecha__lte=fin_mes_pasado)
            elif periodo == 'personalizado':
                fecha_inicio = form.cleaned_data['fecha_inicio']
                fecha_fin = form.cleaned_data['fecha_fin']
                if fecha_inicio and fecha_fin:
                    entradas = entradas.filter(fecha__gte=fecha_inicio, fecha__lte=fecha_fin)
            
            # Aplicar filtros adicionales
            if form.cleaned_data['cliente']:
                entradas = entradas.filter(proyecto__cliente=form.cleaned_data['cliente'])
            
            if form.cleaned_data['proyecto']:
                entradas = entradas.filter(proyecto=form.cleaned_data['proyecto'])
            
            # Ordenar por fecha y hora
            entradas = entradas.order_by('-fecha', '-hora_inicio')
            
            # Calcular totales
            total_horas = sum(entrada.horas for entrada in entradas)
            total_facturado = sum(
                entrada.horas * (entrada.proyecto.tarifa_por_hora or 0) 
                for entrada in entradas
            )
            
            # Calcular estadísticas adicionales
            if entradas:
                # Estadísticas por cliente
                clientes_stats = {}
                for entrada in entradas:
                    cliente = entrada.proyecto.cliente.nombre
                    if cliente not in clientes_stats:
                        clientes_stats[cliente] = 0
                    clientes_stats[cliente] += entrada.horas
                
                # Estadísticas por proyecto
                proyectos_stats = {}
                for entrada in entradas:
                    proyecto = entrada.proyecto.nombre
                    if proyecto not in proyectos_stats:
                        proyectos_stats[proyecto] = 0
                    proyectos_stats[proyecto] += entrada.horas
                
                # Estadísticas por fecha
                fechas_stats = {}
                for entrada in entradas:
                    fecha_str = entrada.fecha.strftime('%Y-%m-%d')
                    if fecha_str not in fechas_stats:
                        fechas_stats[fecha_str] = 0
                    fechas_stats[fecha_str] += entrada.horas
                
                estadisticas = {
                    'clientes': clientes_stats,
                    'proyectos': proyectos_stats,
                    'fechas': fechas_stats,
                    'promedio_diario': total_horas / len(set(entrada.fecha for entrada in entradas)) if entradas else 0,
                    'entrada_mas_larga': max(entradas, key=lambda x: x.horas) if entradas else None,
                    'proyectos_unicos': len(set(entrada.proyecto for entrada in entradas)),
                    'clientes_unicos': len(set(entrada.proyecto.cliente for entrada in entradas)),
                }
    
    context = {
        'form': form,
        'entradas': entradas[:100],  # Limitar a 100 para performance
        'total_horas': total_horas,
        'total_facturado': total_facturado,
        'estadisticas': estadisticas,
        'total_entradas': len(entradas),
    }
    
    return render(request, 'tracker/reportes.html', context)

def exportar_reporte(request):
    """Exportar reporte en diferentes formatos"""
    if request.method != 'POST':
        return redirect('tracker:reportes')
    
    # Recrear los mismos filtros que en la vista de reportes
    form = ReporteForm(request.POST)
    if not form.is_valid():
        messages.error(request, 'Filtros inválidos para exportación')
        return redirect('tracker:reportes')
    
    # Aplicar los mismos filtros
    entradas = EntradaTiempo.objects.all().select_related('proyecto', 'proyecto__cliente').prefetch_related('etiquetas')
    
    # Filtros de fecha (copiado de la vista reportes)
    periodo = form.cleaned_data['periodo']
    hoy = date.today()
    
    if periodo == 'hoy':
        entradas = entradas.filter(fecha=hoy)
    elif periodo == 'ayer':
        ayer = hoy - timedelta(days=1)
        entradas = entradas.filter(fecha=ayer)
    elif periodo == 'esta_semana':
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        entradas = entradas.filter(fecha__gte=inicio_semana)
    elif periodo == 'personalizado':
        fecha_inicio = form.cleaned_data['fecha_inicio']
        fecha_fin = form.cleaned_data['fecha_fin']
        if fecha_inicio and fecha_fin:
            entradas = entradas.filter(fecha__gte=fecha_inicio, fecha__lte=fecha_fin)
    
    # Filtros adicionales
    if form.cleaned_data['cliente']:
        entradas = entradas.filter(proyecto__cliente=form.cleaned_data['cliente'])
    if form.cleaned_data['proyecto']:
        entradas = entradas.filter(proyecto=form.cleaned_data['proyecto'])
    
    entradas = entradas.order_by('-fecha', '-hora_inicio')
    
    # Obtener formato de exportación
    formato = form.cleaned_data['formato_exportacion']
    
    if formato == 'csv':
        return exportar_csv(entradas)
    elif formato == 'excel':
        return exportar_excel(request, entradas)
    elif formato == 'pdf':
        return exportar_pdf(request, entradas)
    else:
        messages.error(request, 'Formato de exportación no válido')
        return redirect('tracker:reportes')

def exportar_csv(entradas):
    """Exportar entradas a CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reporte_tiempo_{date.today()}.csv"'
    response.write('\ufeff')  # BOM para UTF-8
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'Fecha', 'Cliente', 'Proyecto', 'Hora Inicio', 'Hora Fin', 
        'Horas', 'Descripción', 'Etiquetas', 'Facturado'
    ])
    
    # Datos
    for entrada in entradas:
        etiquetas = ', '.join([etiqueta.nombre for etiqueta in entrada.etiquetas.all()])
        writer.writerow([
            entrada.fecha.strftime('%d/%m/%Y'),
            entrada.proyecto.cliente.nombre,
            entrada.proyecto.nombre,
            entrada.hora_inicio.strftime('%H:%M'),
            entrada.hora_fin.strftime('%H:%M') if entrada.hora_fin else '',
            entrada.horas,
            entrada.descripcion,
            etiquetas,
            'Sí' if entrada.facturado else 'No'
        ])
    
    return response

def exportar_excel(request, entradas):
    """Exportar entradas a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        import io
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Tiempo"
        
        # Headers con estilo
        headers = [
            'Fecha', 'Cliente', 'Proyecto', 'Hora Inicio', 'Hora Fin', 
            'Horas', 'Descripción', 'Etiquetas', 'Facturado'
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos
        for row, entrada in enumerate(entradas, 2):
            etiquetas = ', '.join([etiqueta.nombre for etiqueta in entrada.etiquetas.all()])
            
            ws.cell(row=row, column=1, value=entrada.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=entrada.proyecto.cliente.nombre)
            ws.cell(row=row, column=3, value=entrada.proyecto.nombre)
            ws.cell(row=row, column=4, value=entrada.hora_inicio.strftime('%H:%M'))
            ws.cell(row=row, column=5, value=entrada.hora_fin.strftime('%H:%M') if entrada.hora_fin else '')
            ws.cell(row=row, column=6, value=float(entrada.horas))
            ws.cell(row=row, column=7, value=entrada.descripcion)
            ws.cell(row=row, column=8, value=etiquetas)
            ws.cell(row=row, column=9, value='Sí' if entrada.facturado else 'No')
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_tiempo_{date.today()}.xlsx"'
        
        return response
    
    except ImportError as e:
        # Si reportlab no está instalado, usar CSV como fallback
        print(f"Error importando reportlab: {e}")
        messages.warning(request, f'PDF no disponible ({e}), exportando como CSV')
        return exportar_csv(entradas)
    except Exception as e:
        print(f"Error general en exportar_pdf: {e}")
        messages.error(request, f'Error al generar PDF: {e}')
        return exportar_csv(entradas)

@csrf_exempt
def crear_proyecto_ajax(request):
    """Crear proyecto via AJAX desde el modal"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            if not data.get('cliente') or not data.get('nombre'):
                return JsonResponse({
                    'success': False,
                    'error': 'Cliente y nombre son requeridos'
                })
            
            cliente = get_object_or_404(Cliente, pk=data['cliente'])
            
            proyecto = Proyecto.objects.create(
                nombre=data['nombre'],
                cliente=cliente,
                fecha_inicio=datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date(),
                estado='activo'
            )
            
            return JsonResponse({
                'success': True,
                'proyecto': {
                    'id': proyecto.id,
                    'nombre': proyecto.nombre,
                    'cliente_nombre': cliente.nombre,
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def exportar_pdf(request, entradas):
    """Exportar entradas a PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from django.http import HttpResponse
        import io
        
        buffer = io.BytesIO()
        
        # Crear PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
        )
        
        # Título
        title = Paragraph("Reporte de Tiempo", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Información del reporte
        info = Paragraph(f"Generado el: {date.today().strftime('%d/%m/%Y')}", styles['Normal'])
        elements.append(info)
        elements.append(Paragraph(f"Total de entradas: {len(entradas)}", styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Tabla de datos
        data = [['Fecha', 'Cliente', 'Proyecto', 'Horas', 'Descripción']]
        
        for entrada in entradas:
            data.append([
                entrada.fecha.strftime('%d/%m/%Y'),
                entrada.proyecto.cliente.nombre[:20],
                entrada.proyecto.nombre[:25],
                f"{entrada.horas}h",
                entrada.descripcion[:40] + "..." if len(entrada.descripcion) > 40 else entrada.descripcion
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_tiempo_{date.today()}.pdf"'
        
        return response
    
    except ImportError as e:
        # Si reportlab no está instalado, usar CSV como fallback
        print(f"Error importando reportlab: {e}")
        messages.warning(request, f'PDF no disponible ({e}), exportando como CSV')
        return exportar_csv(entradas)